"""Текст из вложений чата и подготовка файлов в workspace."""

from __future__ import annotations

from pathlib import Path
from shutil import copy2
from uuid import uuid4

_TEXT_SUFFIXES = {".txt", ".md", ".markdown"}
_MAX_CHARS = 12_000


def extract_attachment_text(path: str) -> str:
    file_path = Path(path)
    name = file_path.name
    suffix = file_path.suffix.lower()
    if not file_path.is_file():
        return f"файл {name} прикреплён, текст не извлечён (файл не найден)"
    try:
        if suffix in _TEXT_SUFFIXES:
            text = file_path.read_text(encoding="utf-8", errors="replace")
        elif suffix == ".pdf":
            text = _read_pdf(file_path)
        elif suffix == ".docx":
            text = _read_docx(file_path)
        elif suffix == ".doc":
            text = _read_doc(file_path)
        elif suffix == ".xlsx":
            text = _read_xlsx(file_path)
        else:
            text = ""
    except Exception:  # noqa: BLE001
        text = ""
    text = (text or "").strip()
    if not text:
        return f"файл {name} прикреплён, текст не извлечён"
    if len(text) > _MAX_CHARS:
        text = text[:_MAX_CHARS].rstrip() + "\n…"
    return text


def format_attachments_block(paths: list[str]) -> str:
    if not paths:
        return ""
    parts: list[str] = ["", "Прикреплено:"]
    for path in paths:
        name = Path(path).name
        extracted = extract_attachment_text(path)
        parts.append(f"— {name} ({path})")
        parts.append(extracted)
    return "\n".join(parts).rstrip()


def stage_attachments(paths: list[str], workspace_dir: str) -> list[str]:
    if not paths:
        return []
    dest_dir = Path(workspace_dir or ".") / "uploads"
    dest_dir.mkdir(parents=True, exist_ok=True)
    staged: list[str] = []
    for raw in paths:
        src = Path(raw)
        if not src.is_file():
            continue
        dest = dest_dir / src.name
        if dest.exists():
            dest = dest_dir / f"{src.stem}_{uuid4().hex[:6]}{src.suffix}"
        copy2(src, dest)
        staged.append(str(dest.resolve()))
    return staged


def _read_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError:
        return ""
    reader = PdfReader(str(path))
    return "\n\n".join((page.extract_text() or "") for page in reader.pages)


def _read_docx(path: Path) -> str:
    try:
        import docx
    except ImportError:
        return ""
    document = docx.Document(str(path))
    return "\n".join(para.text for para in document.paragraphs)


def _read_doc(path: Path) -> str:
    return ""


def _read_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""
    workbook = load_workbook(str(path), read_only=True, data_only=True)
    try:
        rows: list[str] = []
        for sheet in workbook.worksheets:
            rows.append(f"[{sheet.title}]")
            for row in sheet.iter_rows(values_only=True):
                cells = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
                if cells:
                    rows.append("\t".join(cells))
        return "\n".join(rows)
    finally:
        workbook.close()
