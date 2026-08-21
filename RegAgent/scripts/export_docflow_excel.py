#!/usr/bin/env python3
"""Экспорт Документ.ТД_Поручения в Excel с цветовой маркировкой срочности."""

from __future__ import annotations

import argparse
import re
import sys
from datetime import datetime
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.config import regagent_test_fio, regagent_test_login_enabled
from app.services.docflow_odata import URGENCY_COLORS, handle_docflow_tasks
from app.session_store import saved_fio

HEADERS = [
    "Номер",
    "О чём",
    "Мероприятие",
    "Статус",
    "Срок",
    "Дата документа",
    "Приоритет",
    "Исполнитель",
]

FILL_BY_TIER = {
    tier: PatternFill("solid", fgColor=color.lstrip("#"))
    for tier, color in URGENCY_COLORS.items()
}

TEXT_ON_DARK = Font(color="FFFFFF", bold=False)
TEXT_DEFAULT = Font(color="000000")
HEADER_FILL = PatternFill("solid", fgColor="08745F")
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
PRIORITY_COL = HEADERS.index("Приоритет") + 1
PRIORITY_CRITICAL_FILL = PatternFill("solid", fgColor="FECACA")
PRIORITY_HIGH_FILL = PatternFill("solid", fgColor="FFE0B2")


def _actor_fio() -> str:
    fio = saved_fio()
    if not fio and regagent_test_login_enabled():
        fio = regagent_test_fio()
    return fio


def _tier_font(_tier: str) -> Font:
    return Font(color="1F2937")


def _priority_fill(priority: str) -> PatternFill | None:
    key = re.sub(r"\s+", "", str(priority or "")).casefold()
    if key == "критический":
        return PRIORITY_CRITICAL_FILL
    if key == "высокий":
        return PRIORITY_HIGH_FILL
    return None


def export_excel(*, output: Path, only_open: bool = False, limit: int = 200) -> dict:
    result = handle_docflow_tasks(
        {"only_open": only_open, "limit": limit},
        actor_fio=_actor_fio(),
    )
    tasks = result.get("tasks") or []
    warning = str(result.get("docflow_warning") or "")

    wb = Workbook()
    ws = wb.active
    ws.title = "Поручения"

    if warning:
        ws["A1"] = f"Внимание: {warning}"
        ws["A1"].font = Font(color="DC2626", bold=True)
        ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
        start_row = 3
    else:
        start_row = 1

    for col, title in enumerate(HEADERS, start=1):
        cell = ws.cell(row=start_row, column=col, value=title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    for idx, task in enumerate(tasks, start=1):
        row = start_row + idx
        tier = str(task.get("urgency_tier") or "none")
        fill = FILL_BY_TIER.get(tier, FILL_BY_TIER["none"])
        font = _tier_font(tier)
        values = [
            task.get("number", ""),
            task.get("subject", ""),
            task.get("title", ""),
            task.get("status", ""),
            task.get("due_at", ""),
            task.get("created_at", ""),
            task.get("priority", ""),
            task.get("performer", ""),
        ]
        wrap_cols = {2, 3}
        priority_value = str(task.get("priority", "") or "")
        priority_fill = _priority_fill(priority_value)
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = priority_fill if col == PRIORITY_COL and priority_fill else fill
            cell.font = font
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=col in wrap_cols)

    widths = [14, 36, 42, 12, 18, 18, 12, 28]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = ws.cell(row=start_row + 1, column=1)

    # Легенда
    legend = wb.create_sheet("Легенда")
    legend.append(["Уровень", "Метка", "Цвет", "Условие"])
    legend_rows = [
        ("overdue", "Просрочено", URGENCY_COLORS["overdue"], "срок прошёл"),
        ("due_soon", "Срок 1 день и меньше", URGENCY_COLORS["due_soon"], "срок сегодня или завтра"),
        ("due_3days", "Срок через 3 дня", URGENCY_COLORS["due_3days"], "до срока 2–3 дня"),
        ("accepted", "Принято", URGENCY_COLORS["accepted"], "статус «принято», срок не горит"),
        ("none", "Без срочности", URGENCY_COLORS["none"], "прочие / без срока"),
        ("done_ok", "Выполнено", URGENCY_COLORS["done_ok"], "статус «отменено»"),
        ("priority_critical", "Критический", "#FECACA", "колонка «Приоритет»"),
        ("priority_high", "Высокий", "#FFE0B2", "колонка «Приоритет»"),
    ]
    for row_idx, (tier, label, color, cond) in enumerate(legend_rows, start=2):
        legend.cell(row=row_idx, column=1, value=tier)
        legend.cell(row=row_idx, column=2, value=label)
        c = legend.cell(row=row_idx, column=3, value=color)
        c.fill = PatternFill("solid", fgColor=color.lstrip("#"))
        legend.cell(row=row_idx, column=4, value=cond)

    meta = wb.create_sheet("Сводка")
    meta["A1"] = "Параметр"
    meta["B1"] = "Значение"
    meta["A1"].font = meta["B1"].font = Font(bold=True)
    rows = [
        ("Дата выгрузки", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
        ("Исполнитель", result.get("fio", "")),
        ("Записей", result.get("count", 0)),
        ("Источник", result.get("source", "")),
        ("Только открытые", "нет" if not only_open else "да"),
        ("Предупреждение", warning or "—"),
    ]
    for i, (k, v) in enumerate(rows, start=2):
        meta.cell(row=i, column=1, value=k)
        meta.cell(row=i, column=2, value=v)

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)
    result["output_file"] = str(output)
    return result


def main() -> int:
    parser = argparse.ArgumentParser(description="Экспорт Документ.ТД_Поручения в Excel")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=ROOT / "data" / "exports" / f"porucheniya_{datetime.now():%Y%m%d_%H%M%S}.xlsx",
    )
    parser.add_argument("--only-open", action="store_true", help="Только открытые поручения")
    parser.add_argument("--limit", type=int, default=200)
    args = parser.parse_args()

    result = export_excel(output=args.output, only_open=args.only_open, limit=args.limit)
    print(result.get("summary", ""))
    if result.get("docflow_warning"):
        print("Предупреждение:", result["docflow_warning"])
    print("Файл:", result.get("output_file"))
    return 0 if result.get("count", 0) > 0 else 1


if __name__ == "__main__":
    raise SystemExit(main())
