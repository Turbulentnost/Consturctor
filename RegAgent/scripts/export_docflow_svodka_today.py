#!/usr/bin/env python3
"""Сводка поручений на сегодня: до конца дня и срок < 3 дней."""

from __future__ import annotations

import argparse
import sys
from datetime import date, datetime, timedelta
from pathlib import Path

ROOT = Path(__file__).resolve().parents[1]
SCRIPTS = ROOT / "scripts"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(SCRIPTS) not in sys.path:
    sys.path.insert(0, str(SCRIPTS))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.config import regagent_test_fio, regagent_test_login_enabled
from app.services.docflow_odata import handle_docflow_tasks, sort_tasks
from app.session_store import saved_fio

from export_docflow_excel import (
    BORDER,
    FILL_BY_TIER,
    HEADERS,
    HEADER_FILL,
    HEADER_FONT,
    PRIORITY_COL,
    _actor_fio,
    _priority_fill,
    _tier_font,
)

COL_WIDTHS = [14, 36, 42, 12, 18, 18, 12, 28]
SECTION_FILL = PatternFill("solid", fgColor="E8F5F0")
SECTION_FONT = Font(bold=True, color="06483D", size=12)


def _due_date(task: dict) -> date | None:
    raw = str(task.get("due_at") or "").strip()
    if not raw:
        return None
    try:
        return datetime.fromisoformat(raw.replace(" ", "T")[:19]).date()
    except ValueError:
        return None


def _split_tasks(tasks: list[dict], *, today: date | None = None) -> tuple[list[dict], list[dict]]:
    ref = today or date.today()
    open_tasks = [t for t in tasks if not t.get("done")]
    due_today: list[dict] = []
    due_soon: list[dict] = []
    for task in open_tasks:
        due = _due_date(task)
        if due is None:
            continue
        if due == ref:
            due_today.append(task)
        elif ref < due <= ref + timedelta(days=3):
            due_soon.append(task)
    return sort_tasks(due_today), sort_tasks(due_soon)


def _write_section(
    ws,
    start_row: int,
    title: str,
    tasks: list[dict],
) -> int:
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(HEADERS))
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.fill = SECTION_FILL
    title_cell.font = SECTION_FONT
    title_cell.alignment = Alignment(vertical="center")

    header_row = start_row + 1
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=header_row, column=col, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER

    data_start = header_row + 1
    wrap_cols = {2, 3}
    for offset, task in enumerate(tasks):
        row = data_start + offset
        tier = str(task.get("urgency_tier") or "none")
        row_fill = FILL_BY_TIER.get(tier, FILL_BY_TIER["none"])
        font = _tier_font(tier)
        values = [
            task.get("number", ""),
            task.get("subject", ""),
            task.get("title", ""),
            task.get("status", ""),
            task.get("due_at", ""),
            task.get("created_at", ""),
            task.get("priority", ""),
            task.get("performer", ""),
        ]
        priority_fill = _priority_fill(str(task.get("priority", "") or ""))
        for col, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.fill = priority_fill if col == PRIORITY_COL and priority_fill else row_fill
            cell.font = font
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=col in wrap_cols)

    if not tasks:
        empty_row = data_start
        ws.merge_cells(start_row=empty_row, start_column=1, end_row=empty_row, end_column=len(HEADERS))
        cell = ws.cell(row=empty_row, column=1, value="— нет поручений —")
        cell.alignment = Alignment(horizontal="center")
        return empty_row + 2

    return data_start + len(tasks) + 1


def export_svodka(*, output: Path, limit: int = 400, today: date | None = None) -> dict:
    ref = today or date.today()
    result = handle_docflow_tasks(
        {"only_open": True, "limit": limit},
        actor_fio=_actor_fio(),
    )
    tasks = result.get("tasks") or []
    due_today, due_soon = _split_tasks(tasks, today=ref)

    wb = Workbook()
    ws = wb.active
    ws.title = "Сводка на сегодня"

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(HEADERS))
    top = ws.cell(row=1, column=1, value=f"Сводка на сегодня — {ref.strftime('%d.%m.%Y')}")
    top.font = Font(bold=True, size=14, color="06483D")
    top.alignment = Alignment(horizontal="center")

    row = 3
    row = _write_section(
        ws,
        row,
        f"Закрыть до конца дня ({len(due_today)})",
        due_today,
    )
    row += 1
    _write_section(
        ws,
        row,
        f"Подходящие к сроку — менее 3 дней ({len(due_soon)})",
        due_soon,
    )

    for i, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    meta = wb.create_sheet("Параметры")
    meta.append(["Параметр", "Значение"])
    for key, val in [
        ("Дата сводки", ref.isoformat()),
        ("До конца дня", len(due_today)),
        ("Менее 3 дней", len(due_soon)),
        ("Источник", result.get("source", "")),
        ("Предупреждение", result.get("docflow_warning") or "—"),
    ]:
        meta.append([key, val])

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    summary = {
        **result,
        "output_file": str(output),
        "due_today_count": len(due_today),
        "due_soon_count": len(due_soon),
        "summary": (
            f"Сводка на {ref.strftime('%d.%m.%Y')}: "
            f"сегодня {len(due_today)}, <3 дней {len(due_soon)}"
        ),
    }
    return summary


def main() -> int:
    parser = argparse.ArgumentParser(description="Сводка поручений на сегодня")
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=ROOT / "data" / "exports" / f"svodka_na_segodnya_{date.today():%Y%m%d}.xlsx",
    )
    parser.add_argument("--limit", type=int, default=400)
    args = parser.parse_args()

    result = export_svodka(output=args.output, limit=args.limit)
    print(result.get("summary", ""))
    if result.get("docflow_warning"):
        print("Предупреждение:", result["docflow_warning"])
    print("Файл:", result.get("output_file"))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
