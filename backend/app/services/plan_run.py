"""Запуск агента по правилам из plan_json (не по захардкоженному сценарию)."""

from __future__ import annotations

import os
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Callable
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from app.models.workflow import Workflow
from app.services.workflows.plan_models import PlanRuntime, WorkflowPlan

ProgressCallback = Callable[[str], None]

_TOOLS_ROSELTORG = Path(__file__).resolve().parents[3] / "tools" / "roseltorg_tender_search"
_URL_RE = re.compile(r"https?://[^\s<>\"']+", re.IGNORECASE)

# Map free-text column names from the plan → tender fields.
_COLUMN_ALIASES: dict[str, str] = {
    "название": "title",
    "название тендера": "title",
    "наименование": "title",
    "предмет": "title",
    "цена": "amount",
    "сумма": "amount",
    "нмц": "amount",
    "дата": "deadline",
    "дата окончания": "deadline",
    "срок": "deadline",
    "ссылка": "url",
    "url": "url",
    "ключевые слова": "keywords",
    "ключевые": "keywords",
    "ключи": "keywords",
}


class PlanRunError(RuntimeError):
    pass


@dataclass
class ResolvedRun:
    site_url: str
    keywords: list[str]
    columns: list[str]
    destination: str  # desktop | ...
    export_format: str  # xlsx
    source: str  # runtime | inferred


def desktop_dir() -> Path:
    home = Path.home()
    candidates = [
        home / "Desktop",
        home / "OneDrive" / "Desktop",
        Path(os.environ.get("USERPROFILE", "")) / "Desktop",
        Path(os.environ.get("USERPROFILE", "")) / "OneDrive" / "Desktop",
    ]
    for path in candidates:
        if path and path.is_dir():
            return path
    return home


def plan_dict(workflow: Workflow) -> dict[str, Any]:
    return workflow.plan_json if isinstance(workflow.plan_json, dict) else {}


def resolve_run_spec(workflow: Workflow) -> ResolvedRun | None:
    """Build run rules from structured runtime, else infer from this plan's text only."""
    plan = WorkflowPlan.from_dict(plan_dict(workflow))
    ensure_runtime(plan)
    rt = plan.runtime
    if not rt.kind:
        return None
    if rt.kind != "site_search_excel":
        return None

    keywords = list(rt.keywords)
    if not keywords and rt.keyword_text:
        keywords = _expand_keyword_text(rt.keyword_text)
    if not keywords:
        return None

    columns = list(rt.columns) or ["название", "цена", "дата", "ссылка", "ключевые слова"]
    dest = (rt.export_destination or "desktop").casefold()
    fmt = (rt.export_format or "xlsx").casefold()
    if fmt in {"excel", "xls"}:
        fmt = "xlsx"
    url = (rt.site_url or "").strip()
    if not url:
        url = "https://www.roseltorg.ru/procedures/search"

    return ResolvedRun(
        site_url=url,
        keywords=keywords,
        columns=columns,
        destination=dest,
        export_format=fmt,
        source="runtime" if plan_dict(workflow).get("runtime") else "inferred",
    )


def uses_plan_export(workflow: Workflow) -> bool:
    return resolve_run_spec(workflow) is not None


def ensure_runtime(plan: WorkflowPlan) -> WorkflowPlan:
    """Fill plan.runtime from answers/constraints when planner left it empty."""
    rt = plan.runtime
    if rt.kind and not rt.keywords and rt.keyword_text:
        rt.keywords = _expand_keyword_text(rt.keyword_text)
    if rt.kind and (rt.keywords or rt.keyword_text):
        if not rt.export_format:
            rt.export_format = "xlsx"
        if not rt.export_destination:
            rt.export_destination = "desktop"
        if not rt.columns:
            rt.columns = _extract_columns(_plan_blob(plan)) or [
                "название",
                "цена",
                "дата",
            ]
        return plan
    inferred = infer_runtime(plan)
    if inferred is not None:
        plan.runtime = inferred
    return plan


def infer_runtime(plan: WorkflowPlan) -> PlanRuntime | None:
    """Infer site_search_excel from free-text plan for this agent only."""
    blob = _plan_blob(plan)
    low = blob.casefold()

    wants_excel = "excel" in low or "xlsx" in low or "выгрузк" in low
    wants_desktop = "рабоч" in low and "стол" in low
    has_keywords = "ключев" in low
    if not (wants_excel and (wants_desktop or has_keywords)):
        # Still allow Excel+keywords without explicit desktop.
        if not (wants_excel and has_keywords):
            return None

    url = _extract_url(blob) or ""
    keyword_text = _extract_keyword_block(plan)
    keywords = _expand_keyword_text(keyword_text) if keyword_text else []
    if not keywords:
        return None

    columns = _extract_columns(blob)
    destination = "desktop" if wants_desktop else "desktop"

    return PlanRuntime(
        kind="site_search_excel",
        site_url=url,
        keywords=keywords,
        keyword_text=keyword_text,
        export_format="xlsx",
        export_destination=destination,
        columns=columns,
    )


def _plan_blob(plan: WorkflowPlan) -> str:
    parts = [
        plan.title,
        plan.goal,
        "\n".join(plan.constraints),
        "\n".join(plan.test_criteria),
        "\n".join(plan.out_of_scope),
    ]
    for s in plan.steps:
        parts.append(f"{s.title}\n{s.action}\n{s.done_when}")
    for q in plan.open_questions:
        parts.append(f"{q.question}\n{q.answer}")
    return "\n".join(parts)


def _extract_url(text: str) -> str:
    m = _URL_RE.search(text or "")
    return m.group(0).rstrip(").,;") if m else ""


def _extract_keyword_block(plan: WorkflowPlan) -> str:
    for line in plan.constraints:
        if "ключев" in line.casefold():
            m = re.search(
                r"(?:перечня|слов|фразы)\s*[:：]\s*(.+)$",
                line,
                flags=re.IGNORECASE,
            )
            if m:
                return m.group(1).strip()
            # Drop leading prose up to first ":" if it looks like a list after.
            if ":" in line:
                tail = line.split(":", 1)[1].strip()
                if ";" in tail or "," in tail:
                    return tail
            return line.strip()

    chunks: list[str] = []
    for q in plan.open_questions:
        text = f"{q.question}\n{q.answer}"
        if "ключев" in text.casefold() or "словарь" in text.casefold():
            ans = (q.answer or "").strip()
            if ans:
                chunks.append(ans)
    return "\n".join(chunks)


def _extract_columns(blob: str) -> list[str]:
    low = blob.casefold()
    # Prefer explicit «колонки: a, b, c»
    m = re.search(
        r"колонк[аи]\s*(?:excel[^:]*|выгрузки)?\s*[:—-]\s*([^\n.]+)",
        low,
        flags=re.IGNORECASE,
    )
    found: list[str] = []
    if m:
        raw = m.group(1)
        found = [p.strip(" «»\"'") for p in re.split(r"[,;/]| и ", raw) if p.strip(" «»\"'")]
    # Fallback: look for the classic triad mentioned in answers.
    defaults = []
    for name in ("название", "цена", "дата", "ссылка", "ключевые слова"):
        if name in low:
            defaults.append(name)
    cols = found or defaults or ["название", "цена", "дата"]
    # Always keep link/keywords if plan mentions them and they aren't already there.
    for extra in ("ссылка", "ключевые слова"):
        if extra in low and extra not in cols:
            cols.append(extra)
    # Normalize display labels
    nice = []
    for c in cols:
        c = c.strip().casefold()
        if not c:
            continue
        if c not in nice:
            nice.append(c)
    return nice


def _expand_keyword_text(block: str) -> list[str]:
    block = (block or "").strip()
    if not block:
        return []
    # Prefer semicolon groups (dictionary lines), else newlines.
    if ";" in block:
        raw_parts = [p.strip() for p in block.split(";") if p.strip()]
    else:
        raw_parts = [p.strip() for p in re.split(r"[\n]+", block) if p.strip()]
        if len(raw_parts) == 1 and "," in raw_parts[0]:
            # Single comma list without semicolons.
            raw_parts = [raw_parts[0]]

    path = str(_TOOLS_ROSELTORG)
    if path not in sys.path:
        sys.path.insert(0, path)
    try:
        from roseltorg_tender_search.search_rules import build_queries  # type: ignore

        return build_queries(raw_parts)
    except Exception:
        out: list[str] = []
        seen: set[str] = set()
        for part in raw_parts:
            for chunk in re.split(r"[,/]", part):
                q = chunk.strip()
                if not q:
                    continue
                key = q.casefold()
                if key in seen:
                    continue
                seen.add(key)
                out.append(q)
        return out


def run_site_search_excel(
    workflow: Workflow,
    *,
    on_progress: ProgressCallback | None = None,
    max_queries: int = 80,
) -> dict[str, Any]:
    spec = resolve_run_spec(workflow)
    if spec is None:
        raise PlanRunError("В плане агента нет правил поиска/Excel-выгрузки")

    if spec.export_format != "xlsx":
        raise PlanRunError(f"Формат выгрузки «{spec.export_format}» пока не поддержан")

    queries = spec.keywords[:max_queries]
    if on_progress:
        on_progress(
            f"Правила из паспорта: {len(queries)} ключ., колонки={', '.join(spec.columns)}, "
            f"файл → {spec.destination}, сайт={spec.site_url[:80]}"
        )

    host = (urlparse(spec.site_url).hostname or "").casefold()
    if "roseltorg" not in host and "росэлторг" not in (workflow.title or "").casefold():
        # Generic sites: still try Roseltorg client only for roseltorg; else error clearly.
        if "roseltorg" not in spec.site_url.casefold():
            raise PlanRunError(
                "Поиск с Excel пока реализован для Росэлторг. "
                f"В плане указан другой сайт: {spec.site_url}"
            )

    path = str(_TOOLS_ROSELTORG)
    if path not in sys.path:
        sys.path.insert(0, path)
    try:
        from roseltorg_tender_search import config as rt_config  # type: ignore
        from roseltorg_tender_search.roseltorg_client import search  # type: ignore
    except ImportError as exc:
        raise PlanRunError(
            "Не установлен инструмент поиска (Playwright). "
            "В tools/roseltorg_tender_search: pip install -r requirements.txt "
            "&& python -m playwright install chromium"
        ) from exc

    # Use THIS agent's URL (source filters etc.), not a global hardcode.
    prev_url = rt_config.SEARCH_URL
    try:
        if spec.site_url:
            rt_config.SEARCH_URL = spec.site_url
        def _progress(i: int, total: int, query: str, found: int) -> None:
            if on_progress:
                on_progress(f"Ищу «{query}» ({i}/{total}) — найдено: {found}")

        tenders = search(queries, headless=True, on_progress=_progress)
    finally:
        rt_config.SEARCH_URL = prev_url

    dest_dir = desktop_dir() if spec.destination == "desktop" else desktop_dir()
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    safe_title = re.sub(r"[^\w\-]+", "_", (workflow.title or "agent"), flags=re.UNICODE)[:40]
    dest = dest_dir / f"{safe_title}_{stamp}.xlsx"

    rows = [
        {
            "title": t.title,
            "amount": t.amount,
            "deadline": t.deadline,
            "url": t.url,
            "keywords": ", ".join(t.matched_queries),
        }
        for t in tenders
    ]
    _export_excel(rows, dest, spec.columns)

    if on_progress:
        if not rows:
            on_progress(
                "Поиск завершён без записей. Если Росэлторг блокировал запросы — "
                "повторите запуск; файл Excel всё равно сохранён (пустой)."
            )
        on_progress(f"Excel сохранён: {dest}")

    return {
        "ok": True,
        "file": str(dest),
        "count": len(rows),
        "queries": queries,
        "rows": rows[:30],
        "columns": spec.columns,
        "site_url": spec.site_url,
        "source": spec.source,
    }


def _export_excel(rows: list[dict[str, Any]], dest: Path, columns: list[str]) -> Path:
    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    header_font = Font(bold=True)

    headers = [c[:1].upper() + c[1:] if c else c for c in columns]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 28 if col > 1 else 55
    ws.freeze_panes = "A2"

    field_keys = [_COLUMN_ALIASES.get(c.casefold(), c.casefold()) for c in columns]
    for r_i, row in enumerate(rows, start=2):
        for c_i, key in enumerate(field_keys, start=1):
            val = row.get(key, "")
            ws.cell(row=r_i, column=c_i, value=val).alignment = Alignment(wrap_text=True)

    wb.save(dest)
    return dest


def format_plan_run_answer(result: dict[str, Any]) -> str:
    file_path = result.get("file") or ""
    count = int(result.get("count") or 0)
    cols = ", ".join(result.get("columns") or [])
    lines = [
        f"Готово. Найдено записей: {count}.",
        "Файл Excel сохранён по правилам из паспорта агента:",
        str(file_path),
        f"Колонки: {cols}.",
    ]
    if result.get("site_url"):
        lines.append(f"Источник: {result['site_url']}")
    rows = result.get("rows") or []
    if rows:
        lines.append("")
        lines.append("Примеры:")
        for row in rows[:8]:
            title = str(row.get("title") or "—")
            amount = str(row.get("amount") or "—")
            deadline = str(row.get("deadline") or "—")
            url = str(row.get("url") or "")
            kw = str(row.get("keywords") or "")
            lines.append(f"• {title}")
            lines.append(f"  цена: {amount}; дата: {deadline}")
            if kw:
                lines.append(f"  ключи: {kw}")
            if url:
                lines.append(f"  {url}")
    return "\n".join(lines)
