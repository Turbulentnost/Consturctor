from __future__ import annotations

from pathlib import Path

from app.services.regulation.types import ExtractedBlock, ExtractedDocument, ExtractedTable


def extract_xlsx(path: Path) -> ExtractedDocument:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("Для XLSX требуется зависимость openpyxl") from exc

    workbook = load_workbook(str(path), data_only=True, read_only=True)
    blocks: list[ExtractedBlock] = []

    for page, sheet in enumerate(workbook.worksheets, start=1):
        rows: list[list[str]] = []
        for row in sheet.iter_rows(values_only=True):
            values = ["" if value is None else str(value).strip() for value in row]
            if any(values):
                rows.append(values)
        if not rows:
            continue
        width = max(len(row) for row in rows)
        normalized = [row + [""] * (width - len(row)) for row in rows]
        headers = normalized[0]
        body = normalized[1:] if len(normalized) > 1 else []
        text = "\n".join(" | ".join(cell for cell in row if cell) for row in normalized)
        blocks.append(
            ExtractedBlock(
                page=page,
                section=sheet.title,
                text=text,
                kind="table",
                block_type="table",
                table=ExtractedTable(headers=headers, rows=body),
                confidence=1.0,
            )
        )

    return ExtractedDocument(page_count=len(workbook.worksheets), blocks=blocks)
