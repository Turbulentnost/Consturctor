"""Инструменты работы с Excel в изолированной рабочей папке агента.

Инструменты позволяют агенту создавать, читать и редактировать .xlsx-файлы.
Все файлы лежат в песочнице агента (``<root>/<agent_id>/``); выйти за её
пределы нельзя. Инструменты работают локально (без COM), используя openpyxl.
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from app.tools.ac.tooling import (
    ToolCallResult,
    ToolDefinition,
    ToolExecutionMode,
    ToolSideEffectLevel,
)
from app.tools.ac.agent_workspace import (
    AgentWorkspaceResolver,
    WorkspaceError,
)
from app.tools.ac.base import BaseTool
from app.tools.ac.registry import ToolRegistry


def _cell_value(value: object) -> object:
    """Привести значение ячейки к JSON-совместимому виду."""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _apply_workbook_table_style(
    worksheet,
    *,
    header_rows: int = 1,
    wrap_text_columns: list[int] | None = None,
) -> None:
    """Сетка ячеек и перенос текста — чтобы таблица читалась в Excel."""
    from openpyxl.styles import Alignment, Border, Side

    data_side = Side(style="thin", color="FF90A4AE")
    header_side = Side(style="thin", color="FF546E7A")
    data_border = Border(left=data_side, right=data_side, top=data_side, bottom=data_side)
    header_border = Border(
        left=header_side, right=header_side, top=header_side, bottom=header_side
    )
    wrap_cols = set(wrap_text_columns or [])

    max_row = worksheet.max_row or 1
    max_col = worksheet.max_column or 1
    for row in worksheet.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            is_header = cell.row <= header_rows
            cell.border = header_border if is_header else data_border
            if is_header:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            elif cell.column in wrap_cols:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            else:
                cell.alignment = Alignment(vertical="top")


def _ensure_xlsx(name: str) -> str:
    """Гарантировать расширение .xlsx у имени файла."""
    name = str(name).strip()
    if not name.lower().endswith(".xlsx"):
        name = f"{name}.xlsx"
    return name


def _resolve_workbook_path(workspace, filename: str, *, must_exist: bool = False) -> Path:
    """Путь к .xlsx в out/ или корне рабочей папки агента."""
    name = _ensure_xlsx(filename)
    candidates = [
        workspace.resolve_output(name, must_exist=False),
        workspace.resolve(name, must_exist=False),
    ]
    if must_exist:
        for candidate in candidates:
            if candidate.is_file():
                return candidate
        raise WorkspaceError(f"Файл не найден: {name}")
    return candidates[0]


def _glob_desktop_xlsx(pattern: str) -> list[Path]:
    """Найти .xlsx на рабочем столе Windows (самые новые первыми)."""
    from app.tools.plan_export import desktop_dir

    return sorted(
        desktop_dir().glob(pattern),
        key=lambda item: item.stat().st_mtime,
        reverse=True,
    )


def _resolve_excel_read_path(input_data: dict[str, Any], workspace) -> Path:
    """Прочитать .xlsx из папки агента или с рабочего стола пользователя."""
    from urllib.parse import unquote, urlparse

    from app.tools.plan_export import desktop_dir

    raw = str(
        input_data.get("desktop_path") or input_data.get("absolute_path") or ""
    ).strip()
    if raw:
        if raw.lower().startswith("file:"):
            parsed = urlparse(raw)
            candidate = unquote(parsed.path or "")
            if candidate.startswith("/") and len(candidate) > 2 and candidate[2] == ":":
                candidate = candidate[1:]
        else:
            candidate = raw
        path = Path(candidate).expanduser()
        is_bare_name = not path.is_absolute() and len(path.parts) == 1

        if is_bare_name or "*" in path.name or "?" in path.name:
            pattern = path.name
            matches = _glob_desktop_xlsx(pattern)
            if not matches:
                if "*" in pattern or "?" in pattern:
                    raise WorkspaceError(f"На рабочем столе нет файлов по маске: {pattern}")
                raise WorkspaceError(f"На рабочем столе нет файла: {pattern}")
            path = matches[0].resolve()
        else:
            path = path.resolve()
            if path.suffix.lower() != ".xlsx":
                raise WorkspaceError("Нужен файл .xlsx")
            if not path.is_file():
                raise WorkspaceError(f"Файл не найден: {path}")

        if path.suffix.lower() != ".xlsx":
            raise WorkspaceError("Нужен файл .xlsx")
        if not path.is_file():
            raise WorkspaceError(f"Файл не найден: {path}")
        allowed = {workspace.directory.resolve(), desktop_dir().resolve()}
        if not any(base == path or base in path.parents for base in allowed):
            raise WorkspaceError("Чтение Excel разрешено только с рабочего стола или из папки агента")
        return path

    filename = str(input_data.get("filename") or "").strip()
    if not filename:
        raise WorkspaceError("Укажите filename или desktop_path")
    try:
        return _resolve_workbook_path(workspace, filename, must_exist=True)
    except WorkspaceError:
        pass
    desktop_candidate = desktop_dir() / Path(filename).name
    if desktop_candidate.is_file():
        return desktop_candidate.resolve()
    return workspace.resolve(_ensure_xlsx(filename), must_exist=True)


class _WorkspaceTool(BaseTool):
    """Базовый Excel-инструмент, знающий про рабочую папку агента."""

    def __init__(self, definition: ToolDefinition, resolver: AgentWorkspaceResolver) -> None:
        """Сохранить паспорт и резолвер рабочих папок."""
        super().__init__(definition)
        self._resolver = resolver

    def _workspace(self, input_data: dict):
        """Вернуть рабочую папку агента по runtime_context."""
        agent_id = self._resolver.agent_id_from_input(input_data)
        return self._resolver.for_agent(agent_id)

    def _fail(self, error_type: str, message: str) -> ToolCallResult:
        """Собрать неуспешный результат инструмента."""
        return ToolCallResult(
            ok=False,
            tool_name=self.definition.name,
            error_type=error_type,
            error_message=message,
        )


class ExcelListFilesTool(_WorkspaceTool):
    """Список файлов в рабочей папке агента."""

    def __init__(self, resolver: AgentWorkspaceResolver) -> None:
        """Создать инструмент списка файлов."""
        super().__init__(
            ToolDefinition(
                name="excel.list_files",
                title="Список файлов агента",
                description="Возвращает список файлов в рабочей папке агента.",
                side_effect_level=ToolSideEffectLevel.READ,
                execution_mode=ToolExecutionMode.LOCAL,
                requires_human_approval=False,
                input_schema={"type": "object", "properties": {}},
                output_schema={"type": "object"},
            ),
            resolver,
        )

    def execute(self, input_data: dict) -> ToolCallResult:
        """Вернуть файлы рабочей папки агента."""
        try:
            workspace = self._workspace(input_data)
            files = workspace.list_files()
        except WorkspaceError as exc:
            return self._fail("WORKSPACE_ERROR", str(exc))
        return ToolCallResult(
            ok=True,
            tool_name=self.definition.name,
            output_data={"files": files, "count": len(files)},
        )


class ExcelReadWorkbookTool(_WorkspaceTool):
    """Чтение содержимого .xlsx из рабочей папки агента."""

    def __init__(self, resolver: AgentWorkspaceResolver) -> None:
        """Создать инструмент чтения книги Excel."""
        super().__init__(
            ToolDefinition(
                name="excel.read_workbook",
                title="Чтение Excel",
                description="Читает данные листа .xlsx (заголовки и строки).",
                side_effect_level=ToolSideEffectLevel.READ,
                execution_mode=ToolExecutionMode.LOCAL,
                requires_human_approval=False,
                input_schema={
                    "type": "object",
                    "properties": {
                        "filename": {"type": "string"},
                        "desktop_path": {"type": "string"},
                        "absolute_path": {"type": "string"},
                        "sheet": {"type": "string"},
                        "max_rows": {"type": "integer"},
                    },
                },
                output_schema={"type": "object"},
            ),
            resolver,
        )

    def execute(self, input_data: dict) -> ToolCallResult:
        """Прочитать выбранный лист книги Excel."""
        from openpyxl import load_workbook

        try:
            workspace = self._workspace(input_data)
            path = _resolve_excel_read_path(input_data, workspace)
        except WorkspaceError as exc:
            return self._fail("WORKSPACE_ERROR", str(exc))

        max_rows = int(input_data.get("max_rows") or 500)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001 - openpyxl бросает разные типы
            return self._fail("EXCEL_READ_ERROR", str(exc))

        try:
            sheet_names = list(workbook.sheetnames)
            requested = input_data.get("sheet")
            if requested and requested not in sheet_names:
                return self._fail(
                    "SHEET_NOT_FOUND",
                    f"Лист {requested!r} не найден. Доступны: {sheet_names}",
                )
            worksheet = workbook[requested] if requested else workbook.active
            rows: list[list] = []
            for index, row in enumerate(worksheet.iter_rows(values_only=True)):
                if index >= max_rows:
                    break
                rows.append([_cell_value(cell) for cell in row])
        finally:
            workbook.close()

        return ToolCallResult(
            ok=True,
            tool_name=self.definition.name,
            output_data={
                "filename": path.name,
                "path": str(path),
                "desktop_path": str(path),
                "sheet": worksheet.title,
                "sheets": sheet_names,
                "row_count": len(rows),
                "rows": rows,
                "truncated": len(rows) >= max_rows,
            },
        )


class ExcelCreateWorkbookTool(_WorkspaceTool):
    """Создание нового .xlsx в рабочей папке агента."""

    def __init__(self, resolver: AgentWorkspaceResolver) -> None:
        """Создать инструмент создания книги Excel."""
        super().__init__(
            ToolDefinition(
                name="excel.create_workbook",
                title="Создание Excel",
                description="Создаёт новый .xlsx с заголовками и строками.",
                side_effect_level=ToolSideEffectLevel.CREATE_DRAFT,
                execution_mode=ToolExecutionMode.LOCAL,
                requires_human_approval=False,
                input_schema={
                    "type": "object",
                    "properties": {
                        "filename": {"type": "string"},
                        "sheet": {"type": "string"},
                        "headers": {"type": "array", "items": {"type": "string"}},
                        "rows": {"type": "array"},
                        "row_fills": {
                            "type": "array",
                            "items": {"type": "string"},
                            "description": "ARGB заливка строк данных (без заголовка)",
                        },
                        "header_fill": {
                            "type": "string",
                            "description": "ARGB заливка строки заголовков",
                        },
                        "header_bold": {
                            "type": "boolean",
                            "description": "Жирный шрифт заголовков",
                        },
                        "column_widths": {
                            "type": "array",
                            "items": {"type": "number"},
                            "description": "Ширина колонок",
                        },
                        "freeze_header": {
                            "type": "boolean",
                            "description": "Закрепить строку заголовков",
                        },
                        "cell_borders": {
                            "type": "boolean",
                            "description": "Тонкая сетка по всем ячейкам таблицы",
                        },
                        "wrap_text_columns": {
                            "type": "array",
                            "items": {"type": "integer"},
                            "description": "Номера колонок (1-based) с переносом текста",
                        },
                        "save_to_desktop": {
                            "type": "boolean",
                            "description": "Копия на рабочий стол — только если пользователь явно просил",
                        },
                        "overwrite": {"type": "boolean"},
                    },
                    "required": ["filename"],
                },
                output_schema={"type": "object"},
            ),
            resolver,
        )

    def execute(self, input_data: dict) -> ToolCallResult:
        """Создать новую книгу Excel с данными."""
        from openpyxl import Workbook

        try:
            workspace = self._workspace(input_data)
            path = workspace.resolve_output(_ensure_xlsx(input_data.get("filename", "")))
        except WorkspaceError as exc:
            return self._fail("WORKSPACE_ERROR", str(exc))

        if path.exists() and not input_data.get("overwrite"):
            return self._fail(
                "FILE_EXISTS",
                f"Файл {path.name} уже существует. Передайте overwrite=true "
                "или используйте excel.edit_workbook.",
            )

        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = str(input_data.get("sheet") or "Лист1")[:31]
        headers = input_data.get("headers") or []
        if headers:
            worksheet.append([str(header) for header in headers])
        for row in input_data.get("rows") or []:
            worksheet.append(list(row) if isinstance(row, (list, tuple)) else [row])

        row_fills = input_data.get("row_fills") or []
        if row_fills:
            from openpyxl.styles import PatternFill

            header_offset = 1 if headers else 0
            for idx, fill_argb in enumerate(row_fills):
                if not fill_argb:
                    continue
                row_num = header_offset + idx + 1
                color = str(fill_argb).removeprefix("#").upper()
                if len(color) == 6:
                    color = "FF" + color
                fill = PatternFill(start_color=color[-8:], end_color=color[-8:], fill_type="solid")
                for cell in worksheet[row_num]:
                    cell.fill = fill

        if headers:
            from openpyxl.styles import Font, PatternFill

            header_fill = input_data.get("header_fill")
            header_bold = bool(input_data.get("header_bold"))
            if header_fill or header_bold:
                for cell in worksheet[1]:
                    if header_bold:
                        cell.font = Font(bold=True, color="FFFFFFFF")
                    if header_fill:
                        color = str(header_fill).removeprefix("#").upper()
                        if len(color) == 6:
                            color = "FF" + color
                        cell.fill = PatternFill(
                            start_color=color[-8:],
                            end_color=color[-8:],
                            fill_type="solid",
                        )

        column_widths = input_data.get("column_widths") or []
        if column_widths:
            from openpyxl.utils import get_column_letter

            for index, width in enumerate(column_widths, start=1):
                try:
                    worksheet.column_dimensions[get_column_letter(index)].width = float(width)
                except (TypeError, ValueError):
                    continue

        if input_data.get("freeze_header") and headers:
            worksheet.freeze_panes = "A2"

        if input_data.get("cell_borders") and worksheet.max_row and worksheet.max_column:
            _apply_workbook_table_style(
                worksheet,
                header_rows=1 if headers else 0,
                wrap_text_columns=list(input_data.get("wrap_text_columns") or []),
            )

        desktop_path: Path | None = None
        try:
            workbook.save(path)
            if input_data.get("save_to_desktop"):
                from datetime import datetime

                from app.tools.plan_export import desktop_dir

                import shutil

                desktop_target = desktop_dir() / path.name
                try:
                    shutil.copy2(path, desktop_target)
                    desktop_path = desktop_target
                except OSError as exc:
                    if getattr(exc, "winerror", None) != 32:
                        raise
                    alt = desktop_target.with_stem(
                        f"{desktop_target.stem}_{datetime.now().strftime('%H%M%S')}"
                    )
                    shutil.copy2(path, alt)
                    desktop_path = alt
        except Exception as exc:  # noqa: BLE001
            return self._fail("EXCEL_WRITE_ERROR", str(exc))
        finally:
            workbook.close()

        output: dict[str, Any] = {
            "filename": path.name,
            "sheet": worksheet.title,
            "written_rows": len(input_data.get("rows") or [])
            + (1 if headers else 0),
            "path": str(desktop_path or path),
            "workspace_path": str(path),
        }
        if desktop_path is not None:
            output["desktop_path"] = str(desktop_path)

        return ToolCallResult(
            ok=True,
            tool_name=self.definition.name,
            output_data=output,
        )


class ExcelEditWorkbookTool(_WorkspaceTool):
    """Редактирование существующего .xlsx набором операций."""

    def __init__(self, resolver: AgentWorkspaceResolver) -> None:
        """Создать инструмент редактирования книги Excel."""
        super().__init__(
            ToolDefinition(
                name="excel.edit_workbook",
                title="Редактирование Excel",
                description=(
                    "Изменяет существующий .xlsx: добавляет строки, задаёт ячейки, "
                    "добавляет/удаляет листы."
                ),
                side_effect_level=ToolSideEffectLevel.CREATE_DRAFT,
                execution_mode=ToolExecutionMode.LOCAL,
                requires_human_approval=False,
                input_schema={
                    "type": "object",
                    "properties": {
                        "filename": {"type": "string"},
                        "operations": {"type": "array"},
                    },
                    "required": ["filename", "operations"],
                },
                output_schema={"type": "object"},
            ),
            resolver,
        )

    def execute(self, input_data: dict) -> ToolCallResult:
        """Применить операции редактирования к книге Excel."""
        from openpyxl import load_workbook

        operations = input_data.get("operations")
        if not isinstance(operations, list) or not operations:
            return self._fail(
                "INVALID_OPERATIONS", "Передайте непустой список operations."
            )

        try:
            workspace = self._workspace(input_data)
            path = _resolve_workbook_path(
                workspace,
                str(input_data.get("filename", "")),
                must_exist=True,
            )
        except WorkspaceError as exc:
            return self._fail("WORKSPACE_ERROR", str(exc))

        try:
            workbook = load_workbook(path)
        except Exception as exc:  # noqa: BLE001
            return self._fail("EXCEL_READ_ERROR", str(exc))

        applied: list[str] = []
        try:
            for operation in operations:
                error = self._apply_operation(workbook, operation, applied)
                if error is not None:
                    workbook.close()
                    return self._fail("INVALID_OPERATION", error)
            workbook.save(path)
        except Exception as exc:  # noqa: BLE001
            return self._fail("EXCEL_WRITE_ERROR", str(exc))
        finally:
            workbook.close()

        return ToolCallResult(
            ok=True,
            tool_name=self.definition.name,
            output_data={"filename": path.name, "applied": applied, "path": str(path)},
        )

    def _apply_operation(self, workbook, operation: dict, applied: list[str]) -> str | None:
        """Применить одну операцию; вернуть текст ошибки или None."""
        if not isinstance(operation, dict):
            return "Операция должна быть объектом"
        action = operation.get("action")

        if action == "add_sheet":
            name = str(operation.get("name") or "").strip()
            if not name:
                return "add_sheet требует name"
            workbook.create_sheet(title=name[:31])
            applied.append(f"add_sheet:{name}")
            return None

        if action == "delete_sheet":
            name = str(operation.get("name") or "").strip()
            if name not in workbook.sheetnames:
                return f"delete_sheet: лист {name!r} не найден"
            del workbook[name]
            applied.append(f"delete_sheet:{name}")
            return None

        sheet_name = operation.get("sheet")
        if sheet_name and sheet_name not in workbook.sheetnames:
            return f"Лист {sheet_name!r} не найден"
        worksheet = workbook[sheet_name] if sheet_name else workbook.active

        if action == "append_row":
            values = operation.get("values")
            if not isinstance(values, list):
                return "append_row требует values (список)"
            worksheet.append(values)
            applied.append(f"append_row:{worksheet.title}")
            return None

        if action == "set_cell":
            cell = operation.get("cell")
            if not cell:
                return "set_cell требует cell (например 'B2')"
            worksheet[str(cell)] = operation.get("value")
            applied.append(f"set_cell:{worksheet.title}!{cell}")
            return None

        return f"Неизвестное действие: {action!r}"


def register_excel_tools(
    registry: ToolRegistry,
    resolver: AgentWorkspaceResolver,
    *,
    skip_existing: bool = False,
) -> None:
    """Зарегистрировать Excel-инструменты в реестре."""
    for tool in [
        ExcelListFilesTool(resolver),
        ExcelReadWorkbookTool(resolver),
        ExcelCreateWorkbookTool(resolver),
        ExcelEditWorkbookTool(resolver),
    ]:
        if skip_existing and registry.has_tool(tool.definition.name):
            continue
        registry.register(tool)
