"""Roseltorg search + Excel on the user's Desktop (plan_export tool)."""

from __future__ import annotations

import os
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from app.config import DESKTOP_ROOT, REPO_ROOT
from app.tools.host import ToolHostError

_COLUMN_ALIASES: dict[str, str] = {
    "название": "title",
    "название тендера": "title",
    "наименование": "title",
    "предмет": "title",
    "цена": "amount",
    "сумма": "amount",
    "нмц": "amount",
    "дата": "deadline",
    "дата окончания": "deadline",
    "срок": "deadline",
    "ссылка": "url",
    "url": "url",
    "ключевые слова": "keywords",
    "ключевые": "keywords",
    "ключи": "keywords",
}


def _tools_roseltorg() -> Path:
    for root in (REPO_ROOT / "tools", DESKTOP_ROOT / "tools", DESKTOP_ROOT.parent / "tools"):
        path = root / "roseltorg_tender_search"
        if path.is_dir():
            return path
    return REPO_ROOT / "tools" / "roseltorg_tender_search"


def desktop_dir() -> Path:
    home = Path.home()
    candidates = [
        home / "Desktop",
        home / "OneDrive" / "Desktop",
        Path(os.environ.get("USERPROFILE", "")) / "Desktop",
        Path(os.environ.get("USERPROFILE", "")) / "OneDrive" / "Desktop",
    ]
    for path in candidates:
        if path and path.is_dir():
            return path
    return home


def run_plan_export(arguments: dict[str, Any]) -> dict[str, Any]:
    site_url = str(arguments.get("site_url") or "").strip()
    keywords = [str(k).strip() for k in (arguments.get("keywords") or []) if str(k).strip()]
    columns = [str(c).strip() for c in (arguments.get("columns") or []) if str(c).strip()]
    if not columns:
        columns = ["название", "цена", "дата", "ссылка", "ключевые слова"]
    destination = str(arguments.get("destination") or "desktop").casefold()
    export_format = str(arguments.get("export_format") or "xlsx").casefold()
    workflow_title = str(arguments.get("workflow_title") or "agent")
    source = str(arguments.get("source") or "runtime")
    max_queries = int(arguments.get("max_queries") or 80)

    if export_format not in {"xlsx", "excel", "xls"}:
        raise ToolHostError(f"Формат выгрузки «{export_format}» пока не поддержан")
    if not keywords:
        raise ToolHostError("Нет ключевых слов для поиска")

    queries = keywords[:max_queries]
    if not site_url:
        site_url = "https://www.roseltorg.ru/procedures/search"

    host = (urlparse(site_url).hostname or "").casefold()
    if "roseltorg" not in host and "roseltorg" not in site_url.casefold():
        raise ToolHostError(
            "Поиск с Excel пока реализован для Росэлторг. "
            f"В плане указан другой сайт: {site_url}"
        )

    path = str(_tools_roseltorg())
    if path not in sys.path:
        sys.path.insert(0, path)
    try:
        from roseltorg_tender_search import config as rt_config  # type: ignore
        from roseltorg_tender_search.roseltorg_client import search  # type: ignore
    except ImportError as exc:
        raise ToolHostError(
            "Не установлен инструмент поиска (Playwright). "
            "В tools/roseltorg_tender_search: pip install -r requirements.txt "
            "&& python -m playwright install chromium"
        ) from exc

    prev_url = rt_config.SEARCH_URL
    try:
        rt_config.SEARCH_URL = site_url
        tenders = search(queries, headless=True, on_progress=None)
    finally:
        rt_config.SEARCH_URL = prev_url

    dest_dir = desktop_dir() if destination == "desktop" else desktop_dir()
    stamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    safe_title = re.sub(r"[^\w\-]+", "_", workflow_title, flags=re.UNICODE)[:40]
    dest = dest_dir / f"{safe_title}_{stamp}.xlsx"

    rows = [_tender_to_row(t) for t in _iter_tenders(tenders)]
    _export_excel(rows, dest, columns)

    return {
        "ok": True,
        "file": str(dest),
        "count": len(rows),
        "queries": queries,
        "rows": rows[:30],
        "columns": columns,
        "site_url": site_url,
        "source": source,
    }


def _iter_tenders(items: Any) -> list[Any]:
    if not isinstance(items, list):
        return [items]
    flat: list[Any] = []
    for item in items:
        if isinstance(item, list):
            flat.extend(_iter_tenders(item))
        else:
            flat.append(item)
    return flat


def _tender_to_row(tender: Any) -> dict[str, str]:
    if isinstance(tender, dict):
        title = str(tender.get("title") or tender.get("name") or "Без названия")
        amount = str(tender.get("amount") or tender.get("price") or "")
        deadline = str(tender.get("deadline") or tender.get("date") or "")
        url = str(tender.get("url") or tender.get("link") or "")
        matched = tender.get("matched_queries") or tender.get("keywords") or []
    else:
        title = str(getattr(tender, "title", "") or "Без названия")
        amount = str(getattr(tender, "amount", "") or "")
        deadline = str(getattr(tender, "deadline", "") or "")
        url = str(getattr(tender, "url", "") or "")
        matched = getattr(tender, "matched_queries", []) or []
    if isinstance(matched, str):
        keywords = matched
    else:
        keywords = ", ".join(str(item) for item in matched if str(item).strip())
    return {
        "title": title,
        "amount": amount,
        "deadline": deadline,
        "url": url,
        "keywords": keywords,
    }


def _export_excel(rows: list[dict[str, Any]], dest: Path, columns: list[str]) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ToolHostError(
            "Нужен openpyxl для Excel. Установите: pip install openpyxl"
        ) from exc

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "Результат"
    header_font = Font(bold=True)

    headers = [c[:1].upper() + c[1:] if c else c for c in columns]
    for col, name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = 28 if col > 1 else 55
    ws.freeze_panes = "A2"

    field_keys = [_COLUMN_ALIASES.get(c.casefold(), c.casefold()) for c in columns]
    for r_i, row in enumerate(rows, start=2):
        for c_i, key in enumerate(field_keys, start=1):
            val = row.get(key, "")
            ws.cell(row=r_i, column=c_i, value=val).alignment = Alignment(wrap_text=True)

    wb.save(dest)
    return dest
