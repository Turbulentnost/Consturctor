#!/usr/bin/env python3
"""Перегенерация act_porucheniya_*.xlsx на рабочий стол (построчно по задачам)."""

from __future__ import annotations

import os
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

for line in (ROOT / "infra" / ".env").read_text(encoding="utf-8").splitlines():
    line = line.strip()
    if not line or line.startswith("#") or "=" not in line:
        continue
    key, _, value = line.partition("=")
    os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.services.act_porucheniya_odata import fetch_act_porucheniya_registry
from app.services.act_porucheniya_report import build_act_excel_arguments
from app.services.act_protocol_merge import merge_protocol_documents, parse_protocol_to_documents


def _write_workbook(args: dict, dest: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = str(args.get("sheet") or "Задачи ACT")[:31]
    headers = args.get("headers") or []
    if headers:
        ws.append([str(h) for h in headers])
    for row in args.get("rows") or []:
        ws.append(list(row))

    row_fills = args.get("row_fills") or []
    header_offset = 1 if headers else 0
    for idx, fill_argb in enumerate(row_fills):
        if not fill_argb:
            continue
        color = str(fill_argb).removeprefix("#").upper()
        if len(color) == 6:
            color = "FF" + color
        fill = PatternFill(start_color=color[-8:], end_color=color[-8:], fill_type="solid")
        for cell in ws[header_offset + idx + 1]:
            cell.fill = fill

    if headers:
        hf = args.get("header_fill")
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFFFF")
            if hf:
                color = str(hf).removeprefix("#").upper()
                if len(color) == 6:
                    color = "FF" + color
                cell.fill = PatternFill(start_color=color[-8:], end_color=color[-8:], fill_type="solid")

    for idx, width in enumerate(args.get("column_widths") or [], start=1):
        ws.column_dimensions[chr(64 + idx) if idx <= 26 else "A"].width = float(width)

    if args.get("freeze_header") and headers:
        ws.freeze_panes = "A2"

    dest.parent.mkdir(parents=True, exist_ok=True)
    wb.save(dest)


def main() -> int:
    import argparse

    parser = argparse.ArgumentParser()
    parser.add_argument("workflow_id", nargs="?", default="7e81ded8")
    parser.add_argument("--protocol", default="", help="Путь к txt-протоколу для дополнения")
    args_cli = parser.parse_args()
    workflow_id = args_cli.workflow_id.strip()
    actor = os.environ.get("ERP_LOGIN") or "Жалыбин Максим Дмитриевич"

    payload = fetch_act_porucheniya_registry()
    documents = list(payload.get("documents") or [])
    merge_stats = {}
    if args_cli.protocol:
        proto_path = Path(args_cli.protocol)
        if proto_path.is_file():
            protocol_docs = parse_protocol_to_documents(proto_path.read_text(encoding="utf-8"))
            documents, merge_stats = merge_protocol_documents(documents, protocol_docs)
            print("protocol merge:", merge_stats)
    args = build_act_excel_arguments(
        workflow_id=workflow_id,
        documents=documents,
        actor_fio=actor,
    )
    filename = args["filename"]
    desktop = Path.home() / "Desktop" / filename
    _write_workbook(args, desktop)

    print(f"OK: {desktop}")
    print(f"Документов: {len(documents)}, строк задач: {len(args['rows'])}")
    print("Колонки:", ", ".join(args["headers"]))
    for row in args["rows"]:
        if str(row[0]).startswith("ACT00-00089"):
            print("ACT89 |", " | ".join(str(x) for x in row))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
