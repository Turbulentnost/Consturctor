#!/usr/bin/env python3
"""Проверка агента: правка последней строки в Поручения.xlsx на рабочем столе."""

from __future__ import annotations

import json
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
DESKTOP = ROOT / "desktop"
sys.path.insert(0, str(DESKTOP))

env_path = ROOT / "infra" / ".env"
for line in env_path.read_text(encoding="utf-8").splitlines():
    line = line.strip()
    if not line or line.startswith("#") or "=" not in line:
        continue
    key, _, value = line.partition("=")
    import os

    os.environ.setdefault(key.strip(), value.strip().strip('"').strip("'"))

from app.api_client import ApiClient, ApiError  # noqa: E402

FIO = "Жалыбин Максим Дмитриевич"
PASSWORD = "gm360249"
TASK = (
    "В файле поручения на рабочем столе измени последнюю строчку (новую запись): "
    "номер ACT на порядковый 91 или какой там сейчас последний, и перенеси её наверх."
)


def _ensure_workflow(client: ApiClient) -> str:
    items = client.list_workflows()
    hints = ("act", "аст", "реестр поручений", "act registry", "поручен")
    for item in items:
        title = str(item.title or "").casefold()
        if any(h in title for h in hints):
            return str(item.id or "")

    wf = client.create_workflow(
        notes="ACT registry test workflow for Excel edit",
    )
    client.update_workflow_local_run(
        str(wf.id),
        {
            "seed": "act_porucheniya",
            "tests_status": "pass",
            "can_publish": False,
            "runtime": {"kind": "act_porucheniya"},
        },
    )
    return str(wf.id)


def _snapshot_excel() -> None:
    from openpyxl import load_workbook

    path = Path.home() / "Desktop" / "Поручения.xlsx"
    if not path.is_file():
        print("SNAPSHOT: файл не найден", path)
        return
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    print("SNAPSHOT row2:", rows[1] if len(rows) > 1 else None)
    print("SNAPSHOT row_last:", rows[-1] if rows else None)


def main() -> int:
    _snapshot_excel()
    client = ApiClient()
    print("health:", client.health().status)
    try:
        client.login(FIO, PASSWORD)
    except ApiError as exc:
        print("login failed:", exc)
        return 1

    workflow_id = _ensure_workflow(client)
    print("workflow:", workflow_id)
    print("TASK:", TASK)

    events: list[dict] = []

    def on_event(payload: dict) -> None:
        event_type = str(payload.get("type") or "")
        if event_type in {"status", "agent_message", "error"}:
            text = payload.get("text") or payload.get("message") or ""
            if text:
                print(f"[{event_type}] {text}")
        elif event_type == "tool_call":
            print(f"[tool_call] {payload.get('tool')} args={json.dumps(payload.get('arguments') or {}, ensure_ascii=False)[:400]}")
        elif event_type == "tool_result":
            tool = payload.get("tool")
            result = payload.get("result") or {}
            print(f"[tool_result] {tool}: {json.dumps(result, ensure_ascii=False)[:800]}")
        events.append(payload)

    result = client.stream_workflow_agent_run(
        workflow_id,
        TASK,
        on_event,
        source="script",
        auto_approve=True,
    )
    print("RESULT:", json.dumps(result, ensure_ascii=False)[:1200] if result else "—")
    _snapshot_excel()
    intent = (result or {}).get("intent") or ""
    if intent == "edit_excel":
        return 0
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
