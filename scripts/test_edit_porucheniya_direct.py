#!/usr/bin/env python3
"""Прямой прогон правки Поручения.xlsx: desktop tools + backend logic."""

from __future__ import annotations

import json
import subprocess
import sys
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

ROOT = Path(__file__).resolve().parents[1]
DESKTOP = ROOT / "desktop"
BACKEND = ROOT / "backend"
WORKFLOW_ID = "7e81ded8-5e7b-4f18-ba0c-b9f44bb6ad37"
TARGET = "Поручения.xlsx"
RUNTIME = {"workflow_id": WORKFLOW_ID, "agent_id": WORKFLOW_ID}


def snapshot() -> None:
    from openpyxl import load_workbook

    path = Path.home() / "Desktop" / TARGET
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    print("row2:", rows[1] if len(rows) > 1 else None)
    print("row_last:", rows[-1] if rows else None)


def invoke_desktop_tool(tool: str, arguments: dict) -> dict:
    code = f"""
import json, sys
sys.path.insert(0, {str(DESKTOP)!r})
from app.tools.host import invoke_tool
payload = json.loads({json.dumps(json.dumps(arguments, ensure_ascii=False))!r})
print(json.dumps(invoke_tool({tool!r}, payload), ensure_ascii=False))
"""
    proc = subprocess.run(
        [sys.executable, "-c", code],
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr or proc.stdout)
    return json.loads(proc.stdout.strip())


def build_args(excel_read: dict) -> tuple[dict | None, dict]:
    code = f"""
import json, sys
sys.path.insert(0, {str(BACKEND)!r})
from app.services.act_porucheniya_report import build_act_excel_move_last_row_arguments
payload = json.loads({json.dumps(json.dumps(excel_read, ensure_ascii=False))!r})
args, meta = build_act_excel_move_last_row_arguments(
    payload,
    target_filename={TARGET!r},
    workflow_id={WORKFLOW_ID!r},
    actor_fio="Жалыбин Максим Дмитриевич",
)
print(json.dumps({{"args": args, "meta": meta}}, ensure_ascii=False))
"""
    proc = subprocess.run(
        [sys.executable, "-c", code],
        capture_output=True,
        text=True,
        encoding="utf-8",
    )
    if proc.returncode != 0:
        raise RuntimeError(proc.stderr or proc.stdout)
    payload = json.loads(proc.stdout.strip())
    return payload.get("args"), payload.get("meta") or {}


def main() -> int:
    print("BEFORE")
    snapshot()
    excel_read = invoke_desktop_tool(
        "excel.read_workbook",
        {
            "desktop_path": TARGET,
            "sheet": "Задачи ACT",
            "max_rows": 10000,
            "runtime_context": RUNTIME,
        },
    )
    print("read rows:", excel_read.get("row_count"))
    args, meta = build_args(excel_read)
    if not args:
        print("FAIL build args:", meta)
        return 1
    print("meta:", json.dumps(meta, ensure_ascii=False))
    result = invoke_desktop_tool("excel.create_workbook", {**args, "runtime_context": RUNTIME})
    print("write:", json.dumps(result, ensure_ascii=False)[:500])
    print("AFTER")
    snapshot()
    return 0 if meta.get("new_act") == "ACT00-00091" else 2


if __name__ == "__main__":
    raise SystemExit(main())
