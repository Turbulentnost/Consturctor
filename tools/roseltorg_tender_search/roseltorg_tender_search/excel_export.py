"""Выгрузка результатов в Excel (.xlsx).

Обязательные колонки: «Название тендера», «Сумма», «Дата окончания».
Дополнительно (справочно) — ссылка и совпавшие ключевые слова.
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from .models import Tender

HEADERS = ["Название тендера", "Сумма", "Дата окончания", "Ссылка", "Ключевые слова"]
COLUMN_WIDTHS = [60, 22, 20, 45, 30]


def export(tenders: list[Tender], dest: str | Path) -> Path:
    dest = Path(dest)
    dest.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Тендеры"

    header_font = Font(bold=True)
    for col, name in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=name)
        cell.font = header_font
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(col)].width = COLUMN_WIDTHS[col - 1]
    ws.freeze_panes = "A2"

    for row, t in enumerate(tenders, start=2):
        ws.cell(row=row, column=1, value=t.title).alignment = Alignment(wrap_text=True)
        ws.cell(row=row, column=2, value=t.amount)
        ws.cell(row=row, column=3, value=t.deadline)
        ws.cell(row=row, column=4, value=t.url)
        ws.cell(row=row, column=5, value=", ".join(t.matched_queries))

    wb.save(dest)
    return dest
