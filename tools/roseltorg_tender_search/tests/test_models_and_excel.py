from openpyxl import load_workbook

from roseltorg_tender_search.excel_export import HEADERS, export
from roseltorg_tender_search.models import Tender


def _sample():
    return [
        Tender(
            title="Реконструкция ГРП №5",
            amount="1 200 000,00 руб.",
            deadline="15.09.2026",
            url="https://www.roseltorg.ru/procedures/123",
            procedure_id="123",
            matched_queries=["грп", "реконструкция"],
        ),
        Tender(
            title="Модернизация ГРС",
            amount="5 000 000 руб.",
            deadline="20.09.2026 10:00",
            url="https://www.roseltorg.ru/procedures/456",
            procedure_id="456",
            matched_queries=["модернизация грс"],
        ),
    ]


def test_dedup_key_prefers_id():
    t = Tender(title="x", amount="", deadline="", url="u", procedure_id="42")
    assert t.dedup_key() == "42"
    t2 = Tender(title="Name", amount="", deadline="", url="", procedure_id="")
    assert t2.dedup_key() == "name"


def test_export_creates_valid_xlsx(tmp_path):
    dest = tmp_path / "report.xlsx"
    export(_sample(), dest)
    assert dest.exists()

    wb = load_workbook(dest)
    ws = wb.active
    assert [c.value for c in ws[1]] == HEADERS
    assert ws.cell(row=2, column=1).value == "Реконструкция ГРП №5"
    assert ws.cell(row=2, column=2).value == "1 200 000,00 руб."
    assert ws.cell(row=2, column=3).value == "15.09.2026"
    assert ws.max_row == 3  # заголовок + 2 строки
